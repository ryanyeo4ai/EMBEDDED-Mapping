

from item_object import *
import numpy as np
import matplotlib.pyplot as plt
# from skimage.transform import resize

import pandas as pd
import openpyxl
from PIL import Image as pil_image
from openpyxl.styles import colors
from openpyxl.styles import Color
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from PyPDF2 import PdfFileReader, PdfFileWriter
import os
import sys
import pickle
import csv
from datetime import datetime
import seaborn as sns   

MAIN_FONT_SIZE = 6
MAIN_CELL_WITH = 8
TITLE_FONT_SIZE = 20
SUB_TITLE_FONT_SIZE = 14
SUB_SUB_TITLE_FONT_SIZE = 10

START_MEMORY_USAGE_BLOCK = 5
START_GROUP_BLOCK = 13
START_FREE_BLCOK = 29
START_LOCATE_BLOCK = 36
SHEET_NAME_LENGTH = 20

overange_sheet_name_number = 1
item_name_dict = []
item_name_dict_list= []

PATH_GRAPH = './graph'
PATH_TEMP = './temp'
# OUTPUT_EXCEL = r'c:\\output_excel'
# OUTPUT_PDF = r'c:\\output_pdf'
OUTPUT_EXCEL = r'c:\\Temp\\output_excel'
OUTPUT_PDF = r'c:\\Temp\\output_pdf'
OUTPUT_FINAL = r'c:\\Temp\\output_final'

def resource_path(relative_path): 
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))) 
    return os.path.join(base_path, relative_path)

class Excel:
    def __init__(self, filename):
        
        self.f_name = os.path.splitext(filename)[0]
        self.input_file = self.f_name + '.map'
        self.output_file = self.f_name + '.xlsx'
        # print(f'input_file :{self.input_file}')
        # print(f'output_file :{self.output_file}')
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.font = Font(size=MAIN_FONT_SIZE)
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

    def makeTitle(self, title, lines, tool_name, copmpany_name):
        # print(f'lines : {lines}')
        version = '20220209'
        write_sheet = self.ws
        write_sheet.title = title

        write_sheet.column_dimensions['A'].width = 5
        write_sheet.column_dimensions['B'].width = 12
        write_sheet.column_dimensions['C'].width = 12
        write_sheet.column_dimensions['D'].width = 12
        write_sheet.column_dimensions['E'].width = 12
        write_sheet.column_dimensions['F'].width = 12
        write_sheet.column_dimensions['G'].width = 12
        write_sheet.column_dimensions['H'].width = 2
        write_sheet.column_dimensions['I'].width = 2
        write_sheet.column_dimensions['J'].width = 2
        
        write_sheet.merge_cells('G2:J2')
        img = Image('logo_sm_eng.png')
        write_sheet.add_image(img, 'G2') 
        write_sheet['B2'].alignment = Alignment(
            horizontal='right', vertical='center')
                
        write_sheet.merge_cells('B4:G4')
        dir_name, filename = os.path.split(self.input_file)
        write_sheet['B4'] = "TASKING MAP Parser Report"
        write_sheet['B4'].font = Font(
            size=TITLE_FONT_SIZE, bold=True, color='0000e6')
        write_sheet['B4'].alignment = Alignment(
            horizontal='center', vertical='center')
        
        grayFill = PatternFill(start_color='FF808080',
                               end_color='FF808080',
                               fill_type='solid')
        
        today_str = datetime.today().strftime('%Y/%m/%d')
        thin_border = Border(bottom=Side(style="thin"))
        double_border = Border(bottom=Side(style="double"))
        write_sheet.merge_cells('F5:G5')
        write_sheet['B5'].border = double_border
        write_sheet['C5'].border = double_border
        write_sheet['D5'].border = double_border
        write_sheet['E5'].border = double_border
        write_sheet['F5'].border = double_border
        write_sheet['G5'].border = double_border

        write_sheet['F5'] = 'date : '+ today_str
        write_sheet['F5'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE)
        write_sheet['F5'].alignment = Alignment(
            horizontal='right', vertical='center')
        write_sheet['F5'].border = double_border
        
        write_sheet.merge_cells('B6:G6')    
        write_sheet['B7'] = 'Company : ' + copmpany_name
        write_sheet['B7'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE, bold=True)
        write_sheet['B7'].alignment = Alignment(
            horizontal='left', vertical='center')
        
        write_sheet.merge_cells('B7:G7')    
        write_sheet['B8'] = 'TASKING C Compiler for TriCore/AURIX'
        write_sheet['B8'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE, bold=True)
        write_sheet['B8'].alignment = Alignment(
            horizontal='left', vertical='center')
        
        write_sheet.merge_cells('B8:G8')    
        write_sheet['B9'] = tool_name
        write_sheet['B9'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE, bold=True)
        write_sheet['B9'].alignment = Alignment(
            horizontal='left', vertical='center')
        
        write_sheet.merge_cells('B9:G9')    
        write_sheet['B10'] = 'TASKING MAP Parser Program Version : ' + version
        write_sheet['B10'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE, bold=True)
        write_sheet['B10'].alignment = Alignment(
            horizontal='left', vertical='center')

        write_sheet.merge_cells('B10:G10')
        write_sheet.merge_cells(start_row=11, start_column=2, end_row=12, end_column=7)    
        write_sheet['B11'] = 'MAP file : ' + filename
        write_sheet['B11'].font = Font(size=SUB_SUB_TITLE_FONT_SIZE, bold=True)
        write_sheet['B11'].alignment = Alignment(
            horizontal='left', vertical='center', wrapText=True)
        
        write_sheet['B12'].border = double_border
        write_sheet['C12'].border = double_border
        write_sheet['D12'].border = double_border
        write_sheet['E12'].border = double_border
        write_sheet['F12'].border = double_border
        write_sheet['G12'].border = double_border

        # -------------------------------------                 
        write_sheet['B14'] = 'Summary'
        write_sheet['B14'].font = Font(size=SUB_TITLE_FONT_SIZE)

        write_sheet['B15'] = 'Memory'
        write_sheet['B15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B15'].fill = grayFill
        write_sheet['B15'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['C15'] = 'Code'
        write_sheet['C15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C15'].fill = grayFill
        write_sheet['C15'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['D15'] = 'Date'
        write_sheet['D15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D15'].fill = grayFill
        write_sheet['D15'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['E15'] = 'Reserved'
        write_sheet['E15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E15'].fill = grayFill
        write_sheet['E15'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['F15'] = 'Free'
        write_sheet['F15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['F15'].fill = grayFill
        write_sheet['F15'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['G15'] = 'Total'
        write_sheet['G15'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['G15'].fill = grayFill
        write_sheet['G15'].alignment = Alignment(
            horizontal='center', vertical='center')

        for i in range(15, lines + 16 + 1):
            for j in range(2, 8):
                write_sheet.cell(row=i, column=j).border = self.thin_border

    def make_summary(self, summary_df):
        write_sheet = self.wb['Summary']
        sheet_index = ['B', 'C', 'D', 'E', 'F', 'G']

        # for col in sheet_index:
        #     write_sheet.column_dimensions[col].width = MAIN_CELL_WITH
        sum = [0, 0, 0, 0, 0, 0]

        for i in range(summary_df.shape[0]):
            #print(f'i in summary : {i}')
            if summary_df.iloc[i]['Data'] != None:
                write_sheet[sheet_index[0] +
                            str(i+16)] = summary_df.iloc[i]['Memory']
                write_sheet[sheet_index[0] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                write_sheet[sheet_index[0] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')

                write_sheet[sheet_index[1] + str(i+16)] = '{:#010x}'.format(
                    int(summary_df.iloc[i]['Code'], 16))
                write_sheet[sheet_index[1] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet[sheet_index[1] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                sum[1] += int(summary_df.iloc[i]['Code'], 16)

                write_sheet[sheet_index[2] + str(i+16)] = '{:#010x}'.format(
                    int(summary_df.iloc[i]['Data'], 16))
                write_sheet[sheet_index[2] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet[sheet_index[2] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                sum[2] += int(summary_df.iloc[i]['Data'], 16)

                write_sheet[sheet_index[3] + str(i+16)] = '{:#010x}'.format(
                    int(summary_df.iloc[i]['Reserved'], 16))
                write_sheet[sheet_index[3] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet[sheet_index[3] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                sum[3] += int(summary_df.iloc[i]['Reserved'], 16)

                write_sheet[sheet_index[4] + str(i+16)] = '{:#010x}'.format(
                    int(summary_df.iloc[i]['Free'], 16))
                write_sheet[sheet_index[4] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet[sheet_index[4] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                sum[4] += int(summary_df.iloc[i]['Free'], 16)

                write_sheet[sheet_index[5] + str(i+16)] = '{:#010x}'.format(
                    int(summary_df.iloc[i]['Total'], 16))
                write_sheet[sheet_index[5] +
                            str(i+16)].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet[sheet_index[5] +
                            str(i+16)].font = Font(size=MAIN_FONT_SIZE)
                sum[5] += int(summary_df.iloc[i]['Total'], 16)

        write_sheet[sheet_index[0] + str(summary_df.shape[0]+16)] = 'Total'
        write_sheet[sheet_index[0] + str(summary_df.shape[0]+16)].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet[sheet_index[0] +
                    str(summary_df.shape[0]+16)].font = Font(size=MAIN_FONT_SIZE)
        write_sheet[sheet_index[0] +
                    str(summary_df.shape[0]+16)].fill = PatternFill(fgColor = "FFCC99", fill_type="solid")
        for i in range(len(sum) - 1):
            write_sheet[sheet_index[1+i] +
                        str(summary_df.shape[0]+16)] = '{:#010x}'.format(sum[i+1])
            write_sheet[sheet_index[1+i] +
                        str(summary_df.shape[0]+16)].alignment = Alignment(horizontal='center', vertical='center')
            write_sheet[sheet_index[1+i] +
                        str(summary_df.shape[0]+16)].font = Font(size=MAIN_FONT_SIZE)
            write_sheet[sheet_index[1+i] +
                        str(summary_df.shape[0]+16)].fill = PatternFill(fgColor = "FFCC99", fill_type="solid")

        # summary_df.to_csv('summary.csv', header=None)
        # sum_pickle_file = open('summary.pkl', 'wb')
        # pickle.dump(sum, sum_pickle_file)
        # sum_pickle_file.close()

    def save(self):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))) 
        temp_output_path_file = self.output_file
        file_path = os.path.abspath(temp_output_path_file)
        basename = os.path.basename(file_path)
        name, ext = os.path.splitext(basename)
        output_file = name + '.xlsx'
        # print(f'[12/13] output_file : {OUTPUT_EXCEL + "/" + output_file}')
        
        if not os.path.isdir('OUTPUT/'):
            os.makedirs('OUTPUT/')
            
        # self.wb.save(base_path + "/" + output_file)
        self.wb.save(OUTPUT_EXCEL + "/" + output_file)
        self.wb.save(OUTPUT_FINAL + "/" + output_file)
        self.wb.save('OUTPUT/' + output_file)
        
        
        
    def write_memory(self, item):
        global START_MEMORY_USAGE_BLOCK
        global START_GROUP_BLOCK

        i = 0
        for item_dict in item_name_dict:
            if item_dict == item.item_name:
                    changed_item_name = item_name_dict_list[i]
            i += 1
        # print(f'[12/27] changed_item_name : {changed_item_name}')
        
        write_sheet = self.wb[changed_item_name]
        sheet_index = ['B', 'C', 'D', 'E']

        for col in sheet_index:
            write_sheet.column_dimensions[col].width = MAIN_CELL_WITH

        # print(f'START_MEMORY_USAGE_BLOCK : {START_MEMORY_USAGE_BLOCK}')

        write_sheet['B2'] = item.item_name
        write_sheet['B2'].font = Font(
            size=TITLE_FONT_SIZE, bold=True, color='0000e6')
        write_sheet['B4'] = 'Memory usage in bytes'
        write_sheet['B4'].font = Font(size=20, bold=True)

        grayFill = PatternFill(start_color='FF808080',
                               end_color='FF808080',
                               fill_type='solid')

        write_sheet.merge_cells('C5:E5')
        write_sheet['C5'] = 'Memory usage in bytes'
        write_sheet['C5'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C5'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['C5'].fill = grayFill
        write_sheet['B5'].fill = grayFill
        write_sheet['B6'] = 'Code'
        write_sheet['B6'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B6'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['B6'].fill = grayFill

        write_sheet['B7'] = 'Data'
        write_sheet['B7'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B7'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['B7'].fill = grayFill

        write_sheet['B8'] = 'Reserved'
        write_sheet['B8'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B8'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['B8'].fill = grayFill

        write_sheet['B9'] = 'Free'
        write_sheet['B9'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B9'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['B9'].fill = grayFill

        write_sheet['B10'] = 'Total'
        write_sheet['B10'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B10'].alignment = Alignment(
            horizontal='center', vertical='center')
        write_sheet['B10'].fill = grayFill

        # write_sheet['F6'] = '%'
        # write_sheet['F6'].font = Font(size=MAIN_FONT_SIZE)
        # write_sheet['F7'] = '%'
        # write_sheet['F7'].font = Font(size=MAIN_FONT_SIZE)
        # write_sheet['F8'] = '%'
        # write_sheet['F8'].font = Font(size=MAIN_FONT_SIZE)
        # write_sheet['F9'] = '%'
        # write_sheet['F9'].font = Font(size=MAIN_FONT_SIZE)
        # write_sheet['F10'] = '%'
        # write_sheet['F10'].font = Font(size=MAIN_FONT_SIZE)

        for i in range(START_MEMORY_USAGE_BLOCK, START_MEMORY_USAGE_BLOCK + 5+1):
            for j in range(2, 6):
                write_sheet.cell(row=i, column=j).border = self.thin_border

        START_GROUP_BLOCK = START_MEMORY_USAGE_BLOCK + 5 + 3
        write_sheet['C6'] = '{:#010x}'.format(int(item.addr_code, 16))
        write_sheet['C6'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C6'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['C7'] = '{:#010x}'.format(int(item.addr_data, 16))
        write_sheet['C7'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C7'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['C8'] = '{:#010x}'.format(int(item.addr_reserved, 16))
        write_sheet['C8'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C8'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['C9'] = '{:#010x}'.format(int(item.addr_free, 16))
        write_sheet['C9'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C9'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['C10'] = '{:#010x}'.format(int(item.addr_total, 16))
        write_sheet['C10'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C10'].alignment = Alignment(
            horizontal='center', vertical='center')

        code = int(str(item.addr_code), 16)
        data = int(str(item.addr_data), 16)
        reserved = int(str(item.addr_reserved), 16)
        free = int(str(item.addr_free), 16)
        total = int(str(item.addr_total), 16)

         # 16 digit --> 10 digit
        write_sheet['D6'] = '{:}'.format(code)
        write_sheet['D6'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D6'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['D7'] = '{:}'.format(data)
        write_sheet['D7'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D7'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['D8'] = '{:}'.format(reserved)
        write_sheet['D8'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D8'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['D9'] = '{:}'.format(free)
        write_sheet['D9'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D9'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['D10'] = '{:}'.format(total)
        write_sheet['D10'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D10'].alignment = Alignment(
            horizontal='center', vertical='center')

        # percentage display
        write_sheet['E6'] = '{:.2f}'.format((code/total)*100)  +' %'
        write_sheet['E6'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E6'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['E7'] = '{:.2f}'.format((data/total)*100)  +' %'
        write_sheet['E7'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E7'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['E8'] = '{:.2f}'.format((reserved/total)*100) +' %'
        write_sheet['E8'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E8'].alignment = Alignment(
            horizontal='center', vertical='center')

        
        item_space_file_name = './temp/' + item.item_name + '_space.csv'
        item_space_df = pd.read_csv(item_space_file_name )
        item_counter = item_space_df.shape[0]
        # print(f'[12/13] item_space_df.shape[0] : {item_counter}')
        if item_counter < 1:
            flag_group_data = False
        else:
            flag_group_data = True
                    
        write_sheet['E9'] = '{:.2f}'.format((free/total)*100) +' %'
        write_sheet['E9'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E9'].alignment = Alignment(
            horizontal='center', vertical='center')

        write_sheet['E10'] = '{:.2f}'.format((total/total)*100) +' %'
        write_sheet['E10'].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E10'].alignment = Alignment(
            horizontal='center', vertical='center')

        labels = ['Code', 'Data', 'Reserved', 'Free']
        summary_data = []
        summary_data.append((code/total)*100)
        summary_data.append((data/total)*100)
        summary_data.append((reserved/total)*100)
        summary_data.append((free/total)*100)
        # self.make_pie_plot(item.item_name, summary_data, labels)
        self.make_bar_chart(item.item_name, summary_data, labels)
        img = openpyxl.drawing.image.Image(
            PATH_GRAPH + '/' + item.item_name + '_chart.png')
        img.anchor = 'B12'
        write_sheet.add_image(img)

        return flag_group_data
    
    def write_group(self, item, flag_group_data):
        global START_GROUP_BLOCK
        global START_FREE_BLCOK

        grayFill = PatternFill(start_color='FF808080',
                               end_color='FF808080',
                               fill_type='solid')

        group_data_list = []
        group_label_list = []

        i = 0
        for item_dict in item_name_dict:
            if item_dict == item.item_name:
                    changed_item_name = item_name_dict_list[i]
            i += 1
        # print(f'[12/27] changed_item_name : {changed_item_name}')
        
        write_sheet = self.wb[changed_item_name]
        write_sheet['B22'] = 'Group'
        write_sheet['B22'].font = Font(size=TITLE_FONT_SIZE, bold=True)
        # print(f'[12/13] flag_group_data : {flag_group_data}')
        if flag_group_data == False:
            START_GROUP_BLOCK = 2
            write_sheet['B23'] = 'Group'
            write_sheet['B23'].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B23'].alignment = Alignment(
                horizontal='center', vertical='center')
            write_sheet['B23'].fill = grayFill

            write_sheet.merge_cells('C23:E23')
            write_sheet['C23'] = 'Memory usage in bytes'
            write_sheet['C23'].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C23'].alignment = Alignment(
                horizontal='center', vertical='center')
            write_sheet['C23'].fill = grayFill
            
            write_sheet.merge_cells('B24:E24')
            write_sheet['B24'] = 'Non data'
            write_sheet['B24'].alignment = Alignment(
                horizontal='center', vertical='center')

            for i in range(24, 24 + 1):
                for j in range(2, 6):
                    write_sheet.cell(row=i, column=j).border = self.thin_border
        else:
            START_GROUP_BLOCK = 23
            write_sheet['B23'] = 'Group'
            write_sheet['B23'].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B23'].alignment = Alignment(
                horizontal='center', vertical='center')
            write_sheet['B23'].fill = grayFill

            write_sheet.merge_cells('C23:E23')
            write_sheet['C23'] = 'Memory usage in bytes'
            write_sheet['C23'].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C23'].alignment = Alignment(
                horizontal='center', vertical='center')
            write_sheet['C23'].fill = grayFill

            # print(f'len_group_list : {item.len_group_list}')
            for i in range(START_GROUP_BLOCK, START_GROUP_BLOCK + item.len_group_list + 1 + 1):
                for j in range(2, 6):
                    write_sheet.cell(row=i, column=j).border = self.thin_border

            total = item.total_size
            for i in range(item.len_group_list):
                write_sheet['B' + str(i + START_GROUP_BLOCK + 1)
                            ] = item.group_list[i][0]
                write_sheet['B' + str(i + START_GROUP_BLOCK + 1)
                            ].font = Font(size=MAIN_FONT_SIZE-1)
                write_sheet['B' + str(i + START_GROUP_BLOCK + 1)
                            ].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                #size = int(item.group_list[i][2], 16)
                size = int(item.group_size[item.group_list[i][0]], 16)
                write_sheet['C' + str(i + START_GROUP_BLOCK + 1)
                            ] = '{:#010x}'.format(size)
                write_sheet['C' + str(i + START_GROUP_BLOCK + 1)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['C' + str(i + START_GROUP_BLOCK + 1)
                            ].alignment = Alignment(
                    horizontal='center', vertical='center')
                write_sheet['D' + str(i + START_GROUP_BLOCK + 1)
                            ] = '{:}'.format(size)
                write_sheet['D' + str(i + START_GROUP_BLOCK + 1)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['D' + str(i + START_GROUP_BLOCK + 1)
                            ].alignment = Alignment(
                    horizontal='center', vertical='center')
                temp_percentage_str = '{:.2f}'.format((size/total) * 100 )  
                temp_percentage_str += '%'          
                write_sheet['E' + str(i + START_GROUP_BLOCK + 1)
                            ] = temp_percentage_str
                write_sheet['E' + str(i + START_GROUP_BLOCK + 1)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['E' + str(i + START_GROUP_BLOCK + 1)
                            ].alignment = Alignment(
                    horizontal='center', vertical='center')
                # write_sheet['F' + str(i + START_GROUP_BLOCK + 1)] = '%'
                # write_sheet['F' + str(i + START_GROUP_BLOCK + 1)
                #             ].font = Font(size=MAIN_FONT_SIZE)
                # write_sheet['F' + str(i + START_GROUP_BLOCK + 1)
                #             ].alignment = Alignment(
                #     horizontal='center', vertical='center')
                group_label_list.append(item.group_list[i][0])
                group_data_list.append((size/total) * 100)

            write_sheet['B' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)] = 'Total'
            write_sheet['B' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].alignment = Alignment(
                horizontal='center', vertical='center')
            write_sheet['D' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)] = '{:d}'.format(total)
            write_sheet['D' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['D' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].alignment = Alignment(
                horizontal='center', vertical='center')
            temp_percentage_str = '{:.2f}'.format(100.00 )  
            temp_percentage_str += '%' 
            write_sheet['E' + str(START_GROUP_BLOCK + 1 + item.len_group_list)
                        ] = temp_percentage_str
            write_sheet['E' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['E' + str(START_GROUP_BLOCK + 1 +
                                item.len_group_list)].alignment = Alignment(
                horizontal='center', vertical='center')
            # write_sheet['F' + str(START_GROUP_BLOCK + 1 +
            #                       item.len_group_list)] = '%'
            # write_sheet['F' + str(START_GROUP_BLOCK + 1 +
            #                       item.len_group_list)].font = Font(size=MAIN_FONT_SIZE)
            # write_sheet['F' + str(START_GROUP_BLOCK + 1 +
            #                       item.len_group_list)].alignment = Alignment(
            #     horizontal='center', vertical='center')
            START_BAR_GRAPH = (START_GROUP_BLOCK + 1) + item.len_group_list + 2
            START_FREE_BLCOK = (START_GROUP_BLOCK + 1) + \
                item.len_group_list + 2

            df = pd.DataFrame(
                {'name': group_label_list,
                'data': group_data_list
                },
            )
            df.to_csv(PATH_TEMP + '/' + item.item_name + '_excel_group_name.csv')

            # self.make_horizontal_chart(
            #     item.item_name, group_data_list, group_label_list)

            # img = openpyxl.drawing.image.Image(
            #     PATH_GRAPH + '/' + item.item_name + '_horizontal.png')
            # img.anchor = 'B' + str(START_BAR_GRAPH)
            # # img.anchor = 'G14'
            # write_sheet.add_image(img)

    def write_free(self, item, flag_group_data):
        global START_FREE_BLCOK
        global START_LOCATE_BLOCK

        grayFill = PatternFill(start_color='FF808080',
                               end_color='FF808080',
                               fill_type='solid')

        # print(f'START_FREE_BLCOK : {START_FREE_BLCOK}')
        i = 0
        for item_dict in item_name_dict:
            if item_dict == item.item_name:
                    changed_item_name = item_name_dict_list[i]
            i += 1
        # print(f'[12/27] changed_item_name : {changed_item_name}')
        
        write_sheet = self.wb[changed_item_name]
        write_sheet['B' + str(START_FREE_BLCOK)] = 'Free'

        write_sheet['B' + str(START_FREE_BLCOK)
                    ].font = Font(size=TITLE_FONT_SIZE, bold=True)
        write_sheet['B' + str(1 + START_FREE_BLCOK)] = ''
        write_sheet['B' + str(1 + START_FREE_BLCOK)
                    ].fill = grayFill
        write_sheet.merge_cells(
            'C'+str(START_FREE_BLCOK) + ':D' + str(START_FREE_BLCOK))
        write_sheet['C' + str(1 + START_FREE_BLCOK)] = 'Memory usage in bytes'
        write_sheet['C' + str(1 + START_FREE_BLCOK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C' + str(1 + START_FREE_BLCOK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['C' + str(1 + START_FREE_BLCOK)
                    ].fill = grayFill
        write_sheet['D' + str(1 + START_FREE_BLCOK)
                    ].fill = grayFill

        if flag_group_data == False:
            write_sheet.merge_cells('B31:D31')
            write_sheet['B31'] = 'Non data'
            write_sheet['B31'].alignment = Alignment(
                horizontal='center', vertical='center')

            for i in range(31, 31 + 1):
                for j in range(2, 5):
                    write_sheet.cell(row=i, column=j).border = self.thin_border
        else:
            write_sheet['B' + str(2 + START_FREE_BLCOK)] = 'Total Free'
            write_sheet['B' + str(2 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B' + str(2 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')
            write_sheet['B' + str(2 + START_FREE_BLCOK)
                        ].fill = grayFill
            write_sheet['B' + str(3 + START_FREE_BLCOK)] = 'Gap'
            write_sheet['B' + str(3 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B' + str(3 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')
            write_sheet['B' + str(3 + START_FREE_BLCOK)
                        ].fill = grayFill
            write_sheet['B' + str(4 + START_FREE_BLCOK)] = 'Min Free'
            write_sheet['B' + str(4 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B' + str(4 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')
            write_sheet['B' + str(4 + START_FREE_BLCOK)
                        ].fill = grayFill
            write_sheet['B' + str(5 + START_FREE_BLCOK)] = 'Max Free'
            write_sheet['B' + str(5 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['B' + str(5 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')
            write_sheet['B' + str(5 + START_FREE_BLCOK)
                        ].fill = grayFill
            for i in range(START_FREE_BLCOK + 1, START_FREE_BLCOK + 1 + 5):
                for j in range(2, 5):
                    write_sheet.cell(row=i, column=j).border = self.thin_border

            write_sheet['C' + str(2 + START_FREE_BLCOK)
                        ] = '{:#010x}'.format(item.free)
            write_sheet['C' + str(2 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C' + str(2 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['C' + str(3 + START_FREE_BLCOK)
                        ] = '{:#010x}'.format(item.sum_gap)
            write_sheet['C' + str(3 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C' + str(3 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['C' + str(4 + START_FREE_BLCOK)
                        ] = '{:#010x}'.format(item.min_gap)
            write_sheet['C' + str(4 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C' + str(4 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['C' + str(5 + START_FREE_BLCOK)
                        ] = '{:#010x}'.format(item.max_gap)
            write_sheet['C' + str(5 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['C' + str(5 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['D' + str(2 + START_FREE_BLCOK)
                        ] = '{:}'.format(item.free)
            write_sheet['D' + str(2 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['D' + str(2 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['D' + str(3 + START_FREE_BLCOK)
                        ] = '{:}'.format(item.sum_gap)
            write_sheet['D' + str(3 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['D' + str(3 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['D' + str(4 + START_FREE_BLCOK)
                        ] = '{:}'.format(item.min_gap)
            write_sheet['D' + str(4 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['D' + str(4 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')

            write_sheet['D' + str(5 + START_FREE_BLCOK)
                        ] = '{:}'.format(item.max_gap)
            write_sheet['D' + str(5 + START_FREE_BLCOK)
                        ].font = Font(size=MAIN_FONT_SIZE)
            write_sheet['D' + str(5 + START_FREE_BLCOK)
                        ].alignment = Alignment(horizontal='center', vertical='center')
        START_LOCATE_BLOCK = START_FREE_BLCOK + 4 + 5

    def abnormal_addr_convert(self, data):
        convert_str = ''

        data_1 = int(data, 16)
        mask = int('0x7FFFFFFF', 16)
        data_2 = int('0x70000000', 16)
        data_3 = (data_1 & mask) | data_2

        #convert_str += str(data_3)

        return hex(data_3)

    def preprocess_abnormal_addr(self, item, item_name):
        new_data = []
        converted_data_list = []
        file_name = PATH_TEMP + '/' + item_name + '_locate.csv'
        locate_df = pd.read_csv(file_name)
        locate_list = locate_df.values.tolist()
        for i in range(len(locate_list)):
            new_data = locate_list[i]
            #print(f'[Excel] : {new_data}')
            new_data.append(self.abnormal_addr_convert(locate_list[i][4]))
            converted_data_list.append(new_data)
            #print(f'[Excel] : {new_data}')

        converted_df = pd.DataFrame(data=converted_data_list, columns=[
                                    'WS_1', 'Chip', 'Group', 'Section', 'Size (MAU)', 'Space addr', 'Gap', 'converted'])
        converted_df = converted_df.sort_values(
            by=["converted"], ascending=[True])
        converted_df.reset_index(inplace=True, drop=True)
        converted_df.to_csv(PATH_TEMP + '/' + item_name +
                            '_excel_locate2.csv', mode='w')

    def write_locate(self, item, item_name, flag_group_data):
        global START_LOCATE_BLOCK
        global overange_sheet_name_number 
        
        grayFill = PatternFill(start_color='FF808080',
                               end_color='FF808080',
                               fill_type='solid')

        file_name = PATH_TEMP + '/' + item_name + '_locate.csv'
        # print(f'locate file name : {file_name}')
        locate_df = pd.read_csv(file_name)
        locate_list = locate_df.values.tolist()
        # print(f'locate_list size : {len(locate_list)}')
        # print(f'located_list[0] : {locate_list[0]}')
        file_name = PATH_TEMP + '/' + item_name + '_space.csv'
        # print(f'space file name : {file_name}')

        space_df = pd.read_csv(file_name)
        space_list = space_df.values.tolist()
        # print(f'space_list size : {len(space_list)}')
        # print(f'space_list[0] : {space_list[0]}')

        # print(f'START_LOCATE_BLOCK : {START_LOCATE_BLOCK}')
        

        i = 0
        for item_dict in item_name_dict:
            if item_dict == item.item_name:
                    changed_item_name = item_name_dict_list[i]
            i += 1
        # print(f'[12/27] changed_item_name : {changed_item_name}')
        
        write_sheet = self.wb[changed_item_name]
                
        write_sheet['B' + str(0 + START_LOCATE_BLOCK)] = 'Locate Result'
        write_sheet['B' + str(0 + START_LOCATE_BLOCK)
                    ].font = Font(size=TITLE_FONT_SIZE, bold=True)

        write_sheet['B' + str(1 + START_LOCATE_BLOCK)] = 'Space'
        write_sheet['B' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['B' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['B' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['C' + str(1 + START_LOCATE_BLOCK)] = 'Group'
        write_sheet['C' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['C' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['C' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['D' + str(1 + START_LOCATE_BLOCK)] = 'Section'
        write_sheet['D' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['D' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['D' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['E' + str(1 + START_LOCATE_BLOCK)] = 'Size(MAU)'
        write_sheet['E' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['E' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['E' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['F' + str(1 + START_LOCATE_BLOCK)] = 'Spaceaddr'
        write_sheet['F' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['F' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['F' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['G' + str(1 + START_LOCATE_BLOCK)] = 'Aligment'
        write_sheet['G' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['G' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['G' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill
        write_sheet['H' + str(1 + START_LOCATE_BLOCK)] = 'Gap'
        write_sheet['H' + str(1 + START_LOCATE_BLOCK)
                    ].font = Font(size=MAIN_FONT_SIZE)
        write_sheet['H' + str(1 + START_LOCATE_BLOCK)
                    ].alignment = Alignment(horizontal='center', vertical='center')
        write_sheet['H' + str(1 + START_LOCATE_BLOCK)
                    ].fill = grayFill

        if flag_group_data == False:
            write_sheet.merge_cells('B40:H40')
            write_sheet['B40'] = 'Non data'
            write_sheet['B40'].alignment = Alignment(
                horizontal='center', vertical='center')

            for i in range(40, 40 + 1):
                for j in range(2, 9):
                    write_sheet.cell(row=i, column=j).border = self.thin_border
        else:
            for i in range((1 + START_LOCATE_BLOCK), (1 + START_LOCATE_BLOCK + 2) + len(space_list) - 1):
                for j in range(2, 9):
                    write_sheet.cell(row=i, column=j).border = self.thin_border

            for i in range(len(space_list)):
                # self.find_locate_index(space_list, locate_list, i) # modified by 25_step.py
                locate_index = i
                write_sheet['B' + str(2 + i + START_LOCATE_BLOCK)
                            ] = space_list[i][2]
                write_sheet['B' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE-1)
                write_sheet['B' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                write_sheet['C' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][1]
                write_sheet['C' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE-1)
                write_sheet['C' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(wrap_text=True)
                write_sheet['D' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][2]
                write_sheet['D' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE-1)
                write_sheet['D' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(wrap_text=True)
                write_sheet['E' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][3]
                write_sheet['E' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['E' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet['F' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][4]
                write_sheet['F' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['F' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet['G' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][5]
                write_sheet['G' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['G' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(horizontal='center', vertical='center')
                write_sheet['H' + str(2 + i + START_LOCATE_BLOCK)
                            ] = locate_list[locate_index][7]
                write_sheet['H' + str(2 + i + START_LOCATE_BLOCK)
                            ].font = Font(size=MAIN_FONT_SIZE)
                write_sheet['H' + str(2 + i + START_LOCATE_BLOCK)
                            ].alignment = Alignment(horizontal='center', vertical='center')
                GAP_index_i = 2 + i + START_LOCATE_BLOCK
                write_sheet.cell(GAP_index_i, 8).alignment = Alignment(
                    horizontal='center', vertical='center')

        write_sheet.column_dimensions['A'].width = 1
        write_sheet.column_dimensions['B'].width = 12
        write_sheet.column_dimensions['C'].width = 10
        write_sheet.column_dimensions['D'].width = 20
        write_sheet.column_dimensions['H'].width = 4

    def find_locate_index(self, space_list, locate_list, space_index):
        found_index = -1
        section = space_list[space_index][4]
        for i in range(len(locate_list)):
            if section == locate_list[i][2]:
                found_index = i
                break

        return found_index


    def make_item(self, item):
        global overange_sheet_name_number 
        # if item.item_name == 'dspr0':
        val = item.calculate_group_size()
        # print(f'[12/13] item.calculate_group_size -> {val}')
        if len(item.item_name) > SHEET_NAME_LENGTH:
            changed_item_name = item.item_name[:SHEET_NAME_LENGTH] + ' (skip ' + str(overange_sheet_name_number) + ')'
            # print(f'[12/27] changed_item_name : {changed_item_name}')
            item_name_dict.append(item.item_name)
            item_name_dict_list.append(changed_item_name)
            overange_sheet_name_number += 1
            self.wb.create_sheet(changed_item_name)
        else:
            item_name_dict.append(item.item_name)
            item_name_dict_list.append(item.item_name) 
            self.wb.create_sheet(item.item_name)
        # print(f'[Excel] item_name : {item.item_name}')
        flag_group_data = self.write_memory(item)
        self.write_group(item, flag_group_data)
        self.write_free(item, flag_group_data)
        #self.preprocess_abnormal_addr(item, item.item_name)
        changed_item_name = ''

        self.write_locate(item, item.item_name, flag_group_data)
        # self.write_locate(item)

        return self.ws

    def make_pie_plot(self, item, data, labels):

        try:
            if not os.path.exists(PATH_GRAPH):
                os.makedirs(PATH_GRAPH)
        except OSError:
            print('Error : Failed to create GRAPH directory')

        explode = [0.05, 0.05, 0.05, 0.05]
        
        plt.clf()
        plt.figure(figsize=(2, 1.8))
        # plt.rcParams["font.family"] = "Malgum Gothic"
        plt.rcParams["font.size"] = 6
        # plt.rcParams["figure.figsize"] = (4,4)

        plt.pie(data, labels=labels, autopct='%.1f%%', startangle=260,
                counterclock=False, explode=explode, shadow=True)
        plt.savefig(PATH_GRAPH + '/' + item + '_pie.png')

        pie_img = pil_image.open(PATH_GRAPH + '/' + item + '_pie.png')
        pie_img_resize = pie_img.resize(
            (int(pie_img.width / 3), int(pie_img.height / 3)))
        pie_img_resize.save(PATH_GRAPH + '/' + item + '_res_pie.png')

    def make_stacked_bar_plot(self, item, data, labels):

        with open('data.csv', 'w', newline='') as f:
            write = csv.writer(f)
            write.writerow(labels)
            write.writerow(data)

        # data_df = pd.DataFrame(data)
        # data_df.to_csv('bar_data.csv')
        data_df = pd.read_csv('data.csv')

        plt.clf()
        plt.figure(figsize=(1.2, 1.2))
        plt.rcParams["font.size"] = 10
        # plt.rcParams["figure.figsize"] = (16,10)
        group_data = data
        group_label = labels
        # print(f'[12/10] group_label : {group_label}')
        index = np.arange(len(group_data))
        # print(group_data)
        # plt.barh(index, group_data)
        # plt.bar(stacked=True, index, group_data)
        # plt.bar(index, group_data)
        data_df.plot.bar(stacked=True, rot=0)
        # plt.yticks(index, group_label)
        # plt.legend(group_label, fontsize = 8)
        plt.savefig(PATH_GRAPH + '/' + item + '_bar.png')

        bar_img = pil_image.open(PATH_GRAPH + '/' + item + '_bar.png')
        bar_img_resize = bar_img.resize(
            (int(bar_img.width / 2), int(bar_img.height / 2)))
        bar_img_resize.save(PATH_GRAPH + '/' + item + '_res_bar.png')

    def make_stacked_bar_plot2(self, item, data, labels):

        tuple_data = list(data)
        max_data = len(data)

        # print(f'[12/10] max_data : {max_data}')
        plt.clf()
        plt.figure(figsize=(1.5, 6.0))
        plt.rcParams["font.size"] = 8

        plt.bar('XX', data[0], color="green")
        plt.bar('XX', data[1], color="yellow", bottom=np.array(data[0]))
        plt.bar('XX', data[2], color="red",
                bottom=np.array(data[0]) + np.array(data[1]))
        plt.bar('XX', data[3], color="green", bottom=np.array(
            data[0]) + np.array(data[1]) + np.array(data[2]))
        plt.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
        plt.savefig(PATH_GRAPH + '/' + item + '_bar.png')
    
    
    def make_bar_chart(self, item, data, labels):
        plt.figure(figsize = (4, 2), edgecolor="black")

        #creating the bar plot 
        colors = sns.color_palette('hls',len(data)) ## 색상 지정
        plt.bar(labels, data, width=0.4, color=colors)
        plt.ylabel("Percentage(%)", fontsize=8)
        plt.yticks(np.arange(0, 100,20), fontsize=8)
        plt.title('Memory usage in bytes of '+item, fontsize=10)

        plt.savefig(PATH_GRAPH + '/' + item + '_chart.png')
        chart_img = pil_image.open(PATH_GRAPH + '/' + item + '_chart.png')
        chart_img_resize = chart_img.resize(
            (int(chart_img.width / 3), int(chart_img.height / 3)))
        chart_img_resize.save(PATH_GRAPH + '/' + item + '_res_chart.png')


    def make_horizontal_chart(self, item, data, labels):
        plt.rcParams.update({'figure.autolayout':True})
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.barh(labels, data, height=0.4)
        plt.setp(ax.get_yticklabels(), rotation=45, horizontalalignment='right')
        ax.set(title='Memory usage in bytes of Group in '+item, xlabel='Percentage(%)')
        

        plt.savefig(PATH_GRAPH + '/' + item + '_horizontal.png', transparent=False, dpi=80, bbox_inches="tight")
        hori_img = pil_image.open(PATH_GRAPH + '/' + item + '_horizontal.png')
        # hori_img_resize = hori_img.resize(
        #     (int(hori_img.width / 3), int(hori_img.height / 2)))
        # hori_img_resize.save(PATH_GRAPH + '/' + item + '_res_horizontal.png')



if __name__ == '__main__':
    myExcel = Excel('sample.map')
    myExcel.makeTitle('Summary','')
    myExcel.save()
    # makeTitle()
    # makeSheet("sheet2")
    # writeMemory("sheet2")
    # writeGroup("sheet2")
    # writeFree("sheet2")
    # writeLocate("sheet2")
    # save()
